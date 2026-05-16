"""Sleuth Excel writer.

Generates an xlsx in memory and stages it under a short-lived random
token. The /api/chat/download/{token} endpoint streams it to the browser
exactly once (or until TTL expires, whichever is first) and then drops
the bytes from the in-memory cache.

Why staging instead of returning the bytes inline?

  - Keeps the JSON payload of /api/chat/ask small and snappy. The chat
    bubble appears immediately while the file lives in a background slot.
  - Lets the user re-download by re-clicking the link within the TTL
    without rebuilding the workbook.
  - The cache is per-process and bounded by both row count and TTL, so
    on the 1-vCPU / 2 GB target box we never accumulate spreadsheet bytes
    indefinitely.

Memory budget: each row is roughly 0.5 KB serialized. We cap exports at
5000 rows (executor.py), which is ~2.5 MB per workbook plus openpyxl
overhead — comfortably below the 512 MB container ceiling even with
several recent exports cached.

Concurrency: a `threading.Lock` guards the cache. The chat router is
ASGI but the cache itself is sync; lock overhead is negligible
(microseconds) and avoids subtle races where two requests evict the
same token mid-read.
"""
from __future__ import annotations

import io
import secrets
import threading
import time
from typing import Any, Optional

# openpyxl is the lightest pure-Python xlsx writer that isn't a giant
# dependency. ~3 MB on disk, cold-import ~80 ms — fine on the target box.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - exercised only on broken installs
    OPENPYXL_AVAILABLE = False


class ExcelGenerationError(Exception):
    """Raised when the workbook can't be built (e.g. openpyxl missing)."""


# How long a staged file stays around. Tuned to be long enough that a
# distracted user can still click the link, short enough that we don't
# leak memory across sessions.
_TTL_SECONDS = 30 * 60   # 30 minutes
# Hard cap on the number of staged files at any time. Older entries are
# evicted FIFO-by-creation when this is exceeded.
_MAX_ENTRIES = 50


_cache_lock = threading.Lock()
_cache: dict[str, tuple[bytes, str, float]] = {}
# token -> (xlsx_bytes, filename, expires_at_epoch)


def _evict_expired_locked(now: float) -> None:
    """Drop expired entries. Caller must hold _cache_lock."""
    dead = [tok for tok, (_, _, exp) in _cache.items() if exp <= now]
    for tok in dead:
        _cache.pop(tok, None)


def _evict_oldest_locked() -> None:
    """If the cache is over the size limit, drop the entry with the
    soonest expiry so newer staged files can land. Caller holds the lock."""
    if len(_cache) < _MAX_ENTRIES:
        return
    oldest = min(_cache.items(), key=lambda kv: kv[1][2])
    _cache.pop(oldest[0], None)


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------
_HEADER_STYLE_FILL = "1F2A44"   # Bug Hunter dark accent
_HEADER_STYLE_FG = "FFFFFF"


# Column order — matches what executor._bug_row produces.
_COLUMNS: list[tuple[str, str, int]] = [
    ("id",          "ID",          8),
    ("title",       "Title",       50),
    ("project",     "Project",     20),
    ("status",      "Status",      14),
    ("priority",    "Priority",    12),
    ("environment", "Env",         8),
    ("reporter",    "Reporter",    24),
    ("assignees",   "Assignees",   40),
    ("due_date",    "Due Date",    12),
    ("created_at",  "Created",     22),
    ("updated_at",  "Updated",     22),
]


def _build_workbook(rows: list[dict[str, Any]], description: str) -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise ExcelGenerationError(
            "openpyxl is not installed on the server. "
            "Add it to requirements.txt and redeploy."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Bugs"

    # Top-row banner with the filter description, helpful so the recipient
    # knows what they're looking at without re-asking the chatbot.
    banner = f"Bug Hunter export — {description}" if description else "Bug Hunter export"
    ws.cell(row=1, column=1, value=banner).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))

    # Header row.
    header_fill = PatternFill("solid", fgColor=_HEADER_STYLE_FILL)
    header_font = Font(bold=True, color=_HEADER_STYLE_FG)
    for idx, (_key, header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Data rows.
    for r, row in enumerate(rows, start=3):
        for c, (key, _h, _w) in enumerate(_COLUMNS, start=1):
            val = row.get(key, "")
            # openpyxl chokes on None for some types; coerce to "".
            ws.cell(row=r, column=c, value="" if val is None else val)

    # Freeze the top two rows (banner + header) so scrolling keeps them.
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def stage_workbook(rows: list[dict[str, Any]], filename: str,
                   description: str = "") -> tuple[str, int]:
    """Build the workbook and stash it under a fresh download token.

    Returns (token, size_bytes). Raises ExcelGenerationError on failure.
    """
    payload = _build_workbook(rows, description)
    token = secrets.token_urlsafe(20)
    expires = time.time() + _TTL_SECONDS
    with _cache_lock:
        _evict_expired_locked(time.time())
        _evict_oldest_locked()
        _cache[token] = (payload, filename, expires)
    return token, len(payload)


def fetch_staged(token: str) -> Optional[tuple[bytes, str]]:
    """Return (bytes, filename) if the token is still valid, else None.

    The entry stays in the cache after a fetch — this lets a user click
    the same download link a second time within TTL (e.g. accidental
    close of the download dialog) without re-running the query.
    """
    if not token:
        return None
    now = time.time()
    with _cache_lock:
        _evict_expired_locked(now)
        entry = _cache.get(token)
        if entry is None:
            return None
        payload, filename, expires = entry
        if expires <= now:
            _cache.pop(token, None)
            return None
        return payload, filename


def clear_all_for_test() -> None:
    """Hook for tests — never called from production code paths."""
    with _cache_lock:
        _cache.clear()


__all__ = [
    "ExcelGenerationError",
    "stage_workbook",
    "fetch_staged",
    "OPENPYXL_AVAILABLE",
]
